import ollama
import base64
import json
import re
import os
import fitz  # PyMuPDF
import pandas as pd
from PIL import Image
import io
import time
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

# ══════════════════════════════════════════════════════════════════
#  OLLAMA + MiniCPM-V  —  PDF Table Extractor
#  Works on : Intel Core Ultra 5  |  16 GB RAM  |  No discrete GPU
#  Requirement : Ollama running locally (ollama serve)
# ══════════════════════════════════════════════════════════════════


# ──────────────────────────────────────────────────────────────────
# 0.  CONFIGURATION  (only section you need to edit)
# ──────────────────────────────────────────────────────────────────
PDF_PATH      = "IMAGE_PDF.pdf"          # <- your PDF file path
OUTPUT_FOLDER = "output_tables"          # <- folder for CSVs + Excel
EXCEL_NAME    = "Kharif_Extracted.xlsx"  # <- final Excel file name
OLLAMA_MODEL  = "minicpm-v"              # <- model name in Ollama
DPI           = 200                      # <- render resolution (150-300)
MAX_TOKENS    = 4096                     # <- max tokens per response
TEMPERATURE  = 0.1                       # <- low = more deterministic


# ──────────────────────────────────────────────────────────────────
# 1.  EXTRACTION PROMPT
#     Instructs MiniCPM-V exactly what to extract and how
# ──────────────────────────────────────────────────────────────────
EXTRACTION_PROMPT = """You are a precise agricultural data table extractor.
Carefully look at this document image and extract ALL tables visible.

Return ONLY the following JSON format — no extra text, no markdown:
{
  "tables": [
    {
      "title": "exact table title from the document",
      "headers": ["Column 1", "Column 2", "Column 3"],
      "rows": [
        ["row1_val1", "row1_val2", "row1_val3"],
        ["row2_val1", "row2_val2", "row2_val3"]
      ]
    }
  ]
}

STRICT EXTRACTION RULES:
1.  Preserve ALL numbers exactly as shown
      - Decimals  : 394.28  not  394
      - Negatives : -1.77   not  1.77
      - Zeros     : 0.00    not  blank
2.  Keep row hierarchy intact
      - Main rows : "1. Total Cereals"
      - Sub rows  : "a. Rice", "b. Wheat", "c. Maize"
3.  Multi-level column headers -> join with " > "
      - Example   : "Area Sown > 2024", "Area Sown > 2023"
4.  If a cell spans multiple columns, repeat the value
5.  Include footnotes (lines with * or **) as last rows
6.  Do NOT skip any row — extract every single row
7.  Do NOT add any text outside the JSON block
8.  If no table is found on page return: {"tables": []}
"""


# ──────────────────────────────────────────────────────────────────
# 2.  HELPER UTILITIES
# ──────────────────────────────────────────────────────────────────
def banner(msg: str) -> None:
    """Print a clearly visible section banner."""
    print("\n" + "=" * 64)
    print(f"  {msg}")
    print("=" * 64)


def pil_to_base64(pil_img: Image.Image) -> str:
    """Convert PIL image to base-64 PNG string for Ollama."""
    buffer = io.BytesIO()
    pil_img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


def extract_json(raw_text: str) -> dict | None:
    """
    Robustly extract the first JSON block from a raw LLM response.
    Tries three strategies before giving up.
    """
    # Strategy 1 — response IS clean JSON
    try:
        return json.loads(raw_text.strip())
    except json.JSONDecodeError:
        pass

    # Strategy 2 — find outermost { ... } block
    try:
        start = raw_text.index("{")
        end   = raw_text.rindex("}") + 1
        return json.loads(raw_text[start:end])
    except (ValueError, json.JSONDecodeError):
        pass

    # Strategy 3 — model wrapped JSON in ```json ... ``` fences
    match = re.search(r"```json\s*(.*?)\s*```", raw_text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except json.JSONDecodeError:
            pass

    return None  # all strategies failed


def safe_sheet_name(title: str, index: int) -> str:
    """Build a valid Excel sheet name (max 31 chars, no illegal chars)."""
    name = re.sub(r"[\\/:*?\[\]]", "", str(title))
    name = name.strip()[:24] or f"Table_{index}"
    return f"{index}_{name}"


def check_ollama_running() -> bool:
    """Ping Ollama server — return True if reachable."""
    try:
        ollama.list()
        return True
    except Exception:
        return False


def check_model_available(model_name: str) -> bool:
    """Check whether the model is already pulled in Ollama."""
    try:
        models    = ollama.list()
        available = [m["name"] for m in models.get("models", [])]
        return any(model_name in m for m in available)
    except Exception:
        return False


def print_progress(
    page_num: int, total_pages: int,
    n_tables: int, elapsed_sec: float
) -> None:
    """Print a compact progress bar with ETA."""
    pct       = (page_num / total_pages) * 100
    remaining = (elapsed_sec / page_num) * (total_pages - page_num) \
                if page_num > 0 else 0
    m, s      = int(remaining // 60), int(remaining % 60)
    filled    = int(pct / 5)
    bar       = "#" * filled + "-" * (20 - filled)
    print(f"  [{bar}] {pct:5.1f}%  "
          f"Page {page_num}/{total_pages}  "
          f"Tables so far: {n_tables}  "
          f"ETA: {m}m {s}s")


# ──────────────────────────────────────────────────────────────────
# 3.  PDF -> PAGE IMAGES
# ──────────────────────────────────────────────────────────────────
def pdf_to_page_images(pdf_path: str, dpi: int = 200) -> list:
    """
    Render every PDF page to a PIL Image at the given DPI.
    Higher DPI = better accuracy but more RAM and slower.
    Recommended: 200 DPI for scanned documents.
    """
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    doc    = fitz.open(pdf_path)
    images = []

    print(f"  PDF loaded    : {pdf_path}")
    print(f"  Total pages   : {len(doc)}")
    print(f"  Render DPI    : {dpi}")

    for page_num in range(len(doc)):
        page = doc[page_num]
        pix  = page.get_pixmap(dpi=dpi)
        img  = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        images.append(img)

    doc.close()
    return images


# ──────────────────────────────────────────────────────────────────
# 4.  SINGLE PAGE -> OLLAMA -> PARSED TABLES
# ──────────────────────────────────────────────────────────────────
def extract_tables_from_page(
    page_img   : Image.Image,
    page_num   : int,
    model      : str   = OLLAMA_MODEL,
    max_tokens : int   = MAX_TOKENS,
    temperature: float = TEMPERATURE,
) -> list:
    """
    Send one page image to MiniCPM-V via Ollama.
    Returns a list of table dicts: [{title, headers, rows}, ...]
    Falls back to empty list on any error.
    """
    img_b64 = pil_to_base64(page_img)

    try:
        response = ollama.chat(
            model    = model,
            messages = [{
                "role"   : "user",
                "content": EXTRACTION_PROMPT,
                "images" : [img_b64],
            }],
            options  = {
                "temperature" : temperature,
                "num_predict" : max_tokens,
            },
        )

        raw_text = response["message"]["content"]
        parsed   = extract_json(raw_text)

        if parsed is None:
            print(f"    [!] Page {page_num}: JSON parse failed — saving raw response")
            raw_dir = os.path.join(OUTPUT_FOLDER, "raw_responses")
            os.makedirs(raw_dir, exist_ok=True)
            with open(
                os.path.join(raw_dir, f"page_{page_num}_raw.txt"),
                "w", encoding="utf-8"
            ) as f:
                f.write(raw_text)
            return []

        return parsed.get("tables", [])

    except Exception as e:
        print(f"    [!] Error on page {page_num}: {e}")
        return []


# ──────────────────────────────────────────────────────────────────
# 5.  TABLE DICT -> PANDAS DATAFRAME  (with validation)
# ──────────────────────────────────────────────────────────────────
def table_to_dataframe(table: dict):
    """
    Convert a parsed table dict to a clean DataFrame.
    Pads short rows and trims extra columns automatically.
    """
    headers = table.get("headers", [])
    rows    = table.get("rows",    [])

    if not rows:
        return None

    if headers:
        n_cols     = len(headers)
        normalised = []
        for row in rows:
            if len(row) < n_cols:
                row = list(row) + [""] * (n_cols - len(row))   # pad
            elif len(row) > n_cols:
                row = row[:n_cols]                              # trim
            normalised.append(row)
        df = pd.DataFrame(normalised, columns=headers)
    else:
        df = pd.DataFrame(rows)

    # Strip whitespace from every cell
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    return df


# ──────────────────────────────────────────────────────────────────
# 6.  SAVE RESULTS -> CSV + STYLED EXCEL
# ──────────────────────────────────────────────────────────────────
def save_results(all_results: list, output_folder: str, excel_name: str) -> None:
    """
    For every extracted table:
      - Save an individual CSV  : output_tables/page_X_table_Y.csv
      - Write a styled Excel sheet in the combined workbook
    """
    os.makedirs(output_folder, exist_ok=True)
    excel_path   = os.path.join(output_folder, excel_name)
    total_tables = sum(len(r["tables"]) for r in all_results)

    if total_tables == 0:
        print("  [!] No tables extracted — nothing saved.")
        return

    print(f"  Saving {total_tables} table(s) ...")

    # Excel style constants
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    HDR_FILL  = PatternFill("solid", fgColor="1F5C99")
    DATA_FONT = Font(name="Calibri", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        global_idx = 1

        for result in all_results:
            page_num = result["page"]

            for table in result["tables"]:
                df = table_to_dataframe(table)
                if df is None or df.empty:
                    continue

                title      = table.get("title", f"Table_{global_idx}")
                sheet_name = safe_sheet_name(title, global_idx)

                # -- CSV --
                csv_path = os.path.join(
                    output_folder,
                    f"page_{page_num}_table_{global_idx}.csv"
                )
                df.to_csv(csv_path, index=False, encoding="utf-8-sig")

                # -- Excel sheet --
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                # Style header row
                for cell in ws[1]:
                    cell.font      = HDR_FONT
                    cell.fill      = HDR_FILL
                    cell.alignment = CENTER

                # Style data rows
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = DATA_FONT
                        cell.alignment = LEFT if cell.column == 1 else CENTER

                # Auto column widths
                for col_cells in ws.columns:
                    max_len = max(
                        (len(str(c.value)) for c in col_cells if c.value),
                        default=10
                    )
                    ws.column_dimensions[
                        col_cells[0].column_letter
                    ].width = min(max_len + 4, 50)

                ws.freeze_panes = "A2"  # freeze header row

                print(f"    [OK] Sheet: {sheet_name:30s}  "
                      f"{df.shape[0]:3d} rows x {df.shape[1]:2d} cols  "
                      f"-> {csv_path}")

                global_idx += 1

    print(f"\n  Excel saved: {excel_path}")


# ──────────────────────────────────────────────────────────────────
# 7.  MAIN PIPELINE
# ──────────────────────────────────────────────────────────────────
def main():
    start_time = time.time()

    # ── PRE-FLIGHT CHECKS ─────────────────────────────────────────
    banner("STEP 0 — Pre-flight checks")

    if not check_ollama_running():
        print("  [X] Ollama is NOT running!")
        print("  --> Open a terminal and run:  ollama serve")
        print("  --> Then re-run this script.")
        return
    print("  [OK] Ollama server is running")

    if not check_model_available(OLLAMA_MODEL):
        print(f"  [!] Model '{OLLAMA_MODEL}' not found locally.")
        answer = input(f"  Pull '{OLLAMA_MODEL}' now? (~8 GB download) [y/n]: ")
        if answer.strip().lower() == "y":
            print(f"  Pulling {OLLAMA_MODEL} — please wait ...")
            os.system(f"ollama pull {OLLAMA_MODEL}")
        else:
            print(f"  Run manually:  ollama pull {OLLAMA_MODEL}")
            return
    print(f"  [OK] Model '{OLLAMA_MODEL}' is ready")

    # ── STEP 1 : RENDER PDF PAGES ─────────────────────────────────
    banner("STEP 1 — Rendering PDF to images")
    page_images = pdf_to_page_images(PDF_PATH, dpi=DPI)
    total_pages = len(page_images)

    # ── STEP 2 : EXTRACT TABLES ───────────────────────────────────
    banner("STEP 2 — Extracting tables via MiniCPM-V")
    print(f"  Model       : {OLLAMA_MODEL}")
    print(f"  Temperature : {TEMPERATURE}")
    print(f"  Max tokens  : {MAX_TOKENS}")
    print(f"  Started at  : {datetime.now().strftime('%H:%M:%S')}\n")

    all_results  = []
    total_tables = 0

    for i, page_img in enumerate(page_images, start=1):
        page_start = time.time()
        print(f"\n  -- Page {i}/{total_pages} --")

        tables = extract_tables_from_page(
            page_img    = page_img,
            page_num    = i,
            model       = OLLAMA_MODEL,
            max_tokens  = MAX_TOKENS,
            temperature = TEMPERATURE,
        )

        page_elapsed  = time.time() - page_start
        total_tables += len(tables)

        if tables:
            for t in tables:
                print(f"    [TABLE] '{t.get('title','Untitled')[:50]}'  "
                      f"-> {len(t.get('rows',[]))} rows "
                      f"x {len(t.get('headers',[]))} cols")
        else:
            print("    [INFO] No tables on this page")

        print(f"    [TIME] {page_elapsed:.1f}s")
        all_results.append({"page": i, "tables": tables})

        elapsed = time.time() - start_time
        print_progress(i, total_pages, total_tables, elapsed)

    # ── STEP 3 : SAVE ─────────────────────────────────────────────
    banner("STEP 3 — Saving results")
    save_results(all_results, OUTPUT_FOLDER, EXCEL_NAME)

    # ── SUMMARY ───────────────────────────────────────────────────
    banner("DONE — Summary")
    total_elapsed = time.time() - start_time
    avg_per_page  = total_elapsed / total_pages if total_pages else 0

    print(f"  PDF           : {PDF_PATH}")
    print(f"  Total pages   : {total_pages}")
    print(f"  Tables found  : {total_tables}")
    print(f"  Output folder : {OUTPUT_FOLDER}/")
    print(f"  Excel file    : {OUTPUT_FOLDER}/{EXCEL_NAME}")
    print(f"  Total time    : {int(total_elapsed//60)}m {int(total_elapsed%60)}s")
    print(f"  Avg per page  : {avg_per_page:.1f}s")
    print(f"  Finished at   : {datetime.now().strftime('%H:%M:%S')}")


# ──────────────────────────────────────────────────────────────────
# 8.  ENTRY POINT
# ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()
